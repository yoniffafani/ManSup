from flask import Flask, render_template, request, send_file
import os
import shutil
import easyocr
import pandas as pd
from deep_translator import GoogleTranslator
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageDraw, ImageFont
import tempfile
import warnings
from textwrap import wrap
import socket

# Suppress harmless MPS warnings
warnings.filterwarnings("ignore", message=".*pin_memory.*")

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

reader = easyocr.Reader(['ch_sim', 'en'])  # Chinese simplified + English


# === Utility: clear uploads folder ===
def clear_uploads_folder(folder_path="uploads"):
    """Remove all files and subfolders inside the uploads folder."""
    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f"⚠️ Error deleting {file_path}: {e}")


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    version = request.form.get("version", "desktop")

    if not file or not file.filename.endswith('.xlsx'):
        return "Please upload a valid .xlsx file"

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    wb_in = load_workbook(filepath)
    ws_in = wb_in.active

    results = []
    temp_dir = tempfile.mkdtemp()

    # Create output workbook
    wb_out = Workbook()
    ws_id = wb_out.active
    ws_id.title = "Indonesian"
    ws_en = wb_out.create_sheet("English")
    ws_table = wb_out.create_sheet("Translation Table")

    id_row, en_row = 2, 2
    row_spacing = 60 if version == "mobile" else 60  # consistent spacing

    # Process each image in input Excel
    for idx, image_obj in enumerate(ws_in._images, start=1):
        img = Image.open(BytesIO(image_obj._data()))
        img_path = os.path.join(temp_dir, f"{image_obj.anchor._from.row}_{image_obj.anchor._from.col}.png")
        img.save(img_path)

        # OCR detection
        ocr_results = reader.readtext(img_path, detail=1)
        combined_text = " ".join([res[1] for res in ocr_results])

        # Full text translations
        eng_trans = GoogleTranslator(source='auto', target='en').translate(combined_text) or ""
        indo_trans = GoogleTranslator(source='auto', target='id').translate(combined_text) or ""

        results.append({
            "image_file": os.path.basename(img_path),
            "mandarin_text": combined_text,
            "english": eng_trans,
            "indonesian": indo_trans
        })

        # === Adaptive Small Font Overlay ===
        for lang, translation in [('en', eng_trans), ('id', indo_trans)]:
            img_copy = img.copy()
            draw = ImageDraw.Draw(img_copy)

            for (bbox, mandarin_text) in [(res[0], res[1]) for res in ocr_results]:
                top_left = bbox[0]
                bottom_right = bbox[2]
                x, y = top_left
                box_width = bottom_right[0] - top_left[0]

                # Individual segment translation
                try:
                    translated_segment = GoogleTranslator(source='auto', target=lang).translate(mandarin_text)
                except Exception:
                    translated_segment = ""
                translated_segment = str(translated_segment or "")

                # Start with small readable font and adjust
                font_size = 14
                # --- Cross-platform font path ---
                possible_fonts = [
                    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
                    "/Library/Fonts/Arial.ttf",
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                    "arial.ttf"
                ]
                font_path = next((f for f in possible_fonts if os.path.exists(f)), None)
                if not font_path:
                    font = ImageFont.load_default()
                else:
                    font = ImageFont.truetype(font_path, font_size)

                # Reduce font size until text fits
                while draw.textlength(translated_segment, font=font) > box_width and font_size > 8:
                    font_size -= 1
                    font = ImageFont.truetype(font_path, font_size)

                # Word wrapping based on box width
                max_chars_per_line = max(1, int(box_width / (font_size * 0.6)))
                wrapped_lines = wrap(translated_segment, width=max_chars_per_line)

                # Dynamic text height
                line_height = font.getbbox("A")[3] + 2
                total_height = line_height * len(wrapped_lines)

                # Draw white rectangle behind text
                draw.rectangle(
                    [x, y - total_height - 4, x + box_width, y + 2],
                    fill="white"
                )

                # Draw translated text (magenta)
                for i, line in enumerate(wrapped_lines):
                    draw.text(
                        (x + 2, y - total_height + i * line_height),
                        line,
                        fill=(139, 0, 139),
                        font=font
                    )

            # Save overlay image
            overlay_path = os.path.join(temp_dir, f"overlay_{lang}_{os.path.basename(img_path)}")
            img_copy.save(overlay_path)

            excel_img = ExcelImage(overlay_path)
            if lang == 'id':
                ws_id.add_image(excel_img, f"A{id_row}")
                id_row += row_spacing
            else:
                ws_en.add_image(excel_img, f"A{en_row}")
                en_row += row_spacing

    # === Translation Table Sheet ===
    ws_table.append(["Image File", "Mandarin Text", "English", "Indonesian"])
    for item in results:
        ws_table.append([
            item["image_file"],
            item["mandarin_text"],
            item["english"],
            item["indonesian"]
        ])

    # Save final Excel file
    output_path = os.path.join(RESULT_FOLDER, f'translation_results_{version}.xlsx')
    wb_out.save(output_path)

    # ✅ Clear uploaded files to keep folder clean
    clear_uploads_folder(UPLOAD_FOLDER)

    # Return generated result
    return send_file(output_path, as_attachment=True)


if __name__ == '__main__':
    # Automatically find a free port
    sock = socket.socket()
    sock.bind(('', 0))
    port = sock.getsockname()[1]
    sock.close()

    print(f"\n✅ App running offline on: http://127.0.0.1:{port}\n")
    app.run(host='127.0.0.1', port=port, debug=False)
